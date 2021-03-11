import openpyxl

class ReadExcel:

    @staticmethod
    def readExcelsheet(filelocation,sheetname):

        workbook = openpyxl.load_workbook(filelocation)
        global sheet
        sheet = workbook[sheetname]
        total_row = sheet.max_row
        total_column = sheet.max_column
        return total_row,total_column,sheet